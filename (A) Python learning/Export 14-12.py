import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\Test.xlsx")
sheet = book.active
cell = sheet.cell (row = 4, column= 4)
print(cell.value)

#Writing data to excel

sheet['B3']= "Ajay"
sheet['B7'] = "This is what are you looking for"
book.save("C:\\Users\\Administrator\\Desktop\\Test.xlsx")
book.close()


workbook = openpyxl.Workbook()
ws1 = workbook.create_sheet('ONE')
ws2 = workbook.create_sheet('TWO')

ws1 = workbook.active
ws2 = workbook.active

# Construct test worksheets with some content.
ws1['A1'] = 'ID'
ws1['B1'] = 'Text'
ws1['C1'] = 'Category'

ws1['A2'] = '234'
ws1['B2'] = 'Sample'
ws1['C2'] = 'ASD'

ws2['A1'] = 'ID'
ws2['B1'] = 'Text'
ws2['C1'] = 'Category'

ws2['A2'] = '566'
ws2['B2'] = 'Sample'
ws2['C2'] = 'PED'

ws2['A3'] = '896'
ws2['B3'] = 'SAmple'
ws2['C3'] = 'HEAD'

workbook.save('Ram.xlsx')

